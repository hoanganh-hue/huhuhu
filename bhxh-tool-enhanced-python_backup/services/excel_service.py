"""
Enhanced Excel Service with Batch Processing
"""

import asyncio
import os
import shutil
from datetime import datetime
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from config.config import get_config_instance
from utils.logger import get_logger


class ExcelService:
    """Enhanced Excel Service with Batch Processing"""
    
    def __init__(self):
        self.config = get_config_instance()
        self.logger = get_logger()
        self.pending_writes = []
        self.is_writing = False
        self.batch_buffer = []
    
    async def read_excel_file(self, file_path: Optional[str] = None) -> Dict[str, Any]:
        """Read Excel file with enhanced error handling"""
        target_path = file_path or self.config.files['input_excel']
        
        try:
            self.logger.info(f'📖 Reading Excel file: {target_path}')
            
            # Check if file exists
            if not os.path.exists(target_path):
                raise FileNotFoundError(f'File không tồn tại: {target_path}')
            
            # Check file size
            file_size_mb = os.path.getsize(target_path) / (1024 * 1024)
            
            if file_size_mb > 50:  # 50MB limit
                self.logger.warn(f'⚠️ Large file detected: {file_size_mb:.2f}MB')
            
            # Read workbook
            workbook = load_workbook(target_path, data_only=True)
            sheet_names = workbook.sheetnames
            
            self.logger.info(f'📋 Found {len(sheet_names)} sheet(s): {", ".join(sheet_names)}')
            
            # Process sheets
            result = {}
            
            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                
                # Convert to DataFrame
                df = pd.read_excel(target_path, sheet_name=sheet_name, header=0)
                
                if df.empty:
                    self.logger.warn(f'⚠️ Sheet "{sheet_name}" is empty')
                    continue
                
                # Convert to list of dictionaries
                records = df.to_dict('records')
                
                # Standardize field names
                formatted_data = []
                for index, record in enumerate(records):
                    formatted_record = {}
                    for key, value in record.items():
                        standardized_key = self.standardize_header(key)
                        formatted_record[standardized_key] = self.format_cell_value(value)
                    
                    # Add row metadata
                    formatted_record['_meta'] = {
                        'original_row_index': index + 2,  # +2 because of header and 0-based index
                        'sheet_name': sheet_name
                    }
                    
                    formatted_data.append(formatted_record)
                
                result[sheet_name] = {
                    'headers': list(df.columns),
                    'data': formatted_data,
                    'total_rows': len(formatted_data)
                }
                
                self.logger.info(f'✅ Processed sheet "{sheet_name}": {len(formatted_data)} rows')
            
            total_records = sum(sheet['total_rows'] for sheet in result.values())
            self.logger.info(f'📊 Excel file loaded successfully: {total_records} total records')
            
            return result
            
        except Exception as error:
            self.logger.error(f'❌ Error reading Excel file: {error}', {
                'file_path': target_path,
                'error': str(error)
            })
            raise error
    
    def standardize_header(self, header: str) -> str:
        """Standardize header names to match expected field names"""
        if not header:
            return header
        
        header_str = str(header).lower().strip()
        
        # Header mapping for real-world data
        header_map = {
            'số điện thoại': 'soDienThoai',
            'sđt': 'soDienThoai',
            'sdt': 'soDienThoai',
            'số cccd': 'soCCCD',
            'cccd': 'soCCCD',
            'số cmnd': 'soCCCD',
            'cmnd': 'soCCCD',
            'cmt': 'soCCCD',
            'số cmt/thẻ căn cước': 'soCCCD',  # Thêm biến thể mới
            'họ và tên': 'hoVaTen',
            'họ tên': 'hoVaTen',
            'tên': 'hoVaTen',
            'họ và tên ': 'hoVaTen',
            'địa chỉ': 'diaChi',
            'ngày tháng năm sinh': 'ngayThangNamSinh',
            'ngày sinh': 'ngayThangNamSinh',
            'mã bhxh': 'maBHXH',
            'bhxh': 'maBHXH'
        }
        
        return header_map.get(header_str, header)
    
    def format_cell_value(self, value: Any) -> str:
        """Format cell value for consistency"""
        if pd.isna(value):
            return ''
        
        # Convert to string and trim
        formatted_value = str(value).strip()
        
        # Remove extra quotes that might be added by Excel
        if formatted_value.startswith("'") and len(formatted_value) > 1:
            formatted_value = formatted_value[1:]
        
        return formatted_value
    
    def create_output_structure(self) -> Dict[str, Any]:
        """Create output Excel structure"""
        return {
            'headers': [
                'Số Điện Thoại',
                'Số CCCD',
                'Họ và Tên',
                'Địa Chỉ',
                'Ngày Tháng Năm Sinh',
                'Mã BHXH',
                'Giới Tính',
                'Trạng Thái BHXH',
                'Số Kết Quả',
                'Thời Gian Xử Lý',
                'Trạng Thái Xử Lý'
            ],
            'data': []
        }
    
    def format_output_record(self, input_record: Dict[str, Any], bhxh_result: Dict[str, Any], 
                           processing_time: int) -> List[Any]:
        """Format output record for Excel"""
        output_record = []
        
        # Basic input fields (without extra quotes)
        output_record.append(input_record.get('soDienThoai', ''))
        output_record.append(input_record.get('soCCCD', ''))
        output_record.append(input_record.get('hoVaTen', ''))
        output_record.append(input_record.get('diaChi', ''))
        
        # BHXH results
        if bhxh_result.get('status') == 'success' and bhxh_result.get('soKetQua', 0) > 0:
            first_result = bhxh_result.get('thongTinHoGiaDinh', [{}])[0]
            output_record.append(first_result.get('ngaySinh', ''))
            output_record.append(first_result.get('maBHXH', ''))
            output_record.append(first_result.get('gioiTinh', ''))
            output_record.append(first_result.get('trangThai', ''))
            output_record.append(bhxh_result.get('soKetQua', 0))
        else:
            output_record.append('')  # Ngày sinh
            output_record.append('')  # Mã BHXH - keep empty on error/no data
            output_record.append('')  # Giới tính
            # Put human-readable status or error message in Trạng Thái BHXH
            status_text = (bhxh_result.get('message', 'Lỗi xử lý') 
                          if bhxh_result.get('status') == 'error' 
                          else 'Không tham gia BHXH')
            output_record.append(status_text)
            output_record.append(bhxh_result.get('soKetQua', 0))
        
        # Processing metadata
        output_record.append(f'{processing_time}ms' if processing_time else '')
        output_record.append('Thành công' if bhxh_result.get('status') == 'success' else 'Thất bại')
        
        return output_record
    
    async def add_to_batch(self, input_record: Dict[str, Any], bhxh_result: Dict[str, Any], 
                          processing_time: int):
        """Add record to batch buffer"""
        formatted_record = self.format_output_record(input_record, bhxh_result, processing_time)
        self.batch_buffer.append(formatted_record)
        
        self.logger.debug(f'📝 Added record to batch buffer ({len(self.batch_buffer)}/{self.config.processing["batch_write_size"]})')
        
        # Auto-flush if batch is full
        if len(self.batch_buffer) >= self.config.processing['batch_write_size']:
            await self.flush_batch()
    
    async def flush_batch(self, force: bool = False):
        """Flush batch buffer to Excel"""
        if not self.batch_buffer:
            return
        
        if not force and len(self.batch_buffer) < self.config.processing['batch_write_size']:
            return
        
        try:
            self.logger.info(f'💾 Flushing batch: {len(self.batch_buffer)} records')
            
            await self.write_batch_to_excel(self.batch_buffer.copy())
            self.batch_buffer.clear()  # Clear buffer
            
            self.logger.debug('✅ Batch flushed successfully')
            
        except Exception as error:
            self.logger.error(f'❌ Error flushing batch: {error}')
            raise error
    
    async def write_batch_to_excel(self, records: List[List[Any]], retry_count: int = 0):
        """Write batch of records to Excel"""
        max_retries = 3
        retry_delay = 2000 * (2 ** retry_count)  # Exponential backoff
        
        try:
            # Prevent concurrent writes
            while self.is_writing:
                await asyncio.sleep(0.1)
            
            self.is_writing = True
            
            output_path = self.config.files['output_excel']
            existing_data = []
            output_structure = None
            
            # Read existing file if it exists
            if os.path.exists(output_path):
                try:
                    workbook = load_workbook(output_path)
                    worksheet = workbook.active
                    
                    # Convert to list of lists
                    for row in worksheet.iter_rows(values_only=True):
                        existing_data.append(list(row))
                    
                    # Skip header if exists
                    if existing_data and self.is_header_row(existing_data[0]):
                        output_structure = {
                            'headers': existing_data[0],
                            'data': existing_data[1:]
                        }
                    else:
                        output_structure = self.create_output_structure()
                except Exception as read_error:
                    self.logger.warn(f'⚠️ Error reading existing file, creating new: {read_error}')
                    output_structure = self.create_output_structure()
            else:
                output_structure = self.create_output_structure()
            
            # Add new records
            output_structure['data'].extend(records)
            
            # Create new workbook
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'BHXH Results'
            
            # Write headers
            for col_idx, header in enumerate(output_structure['headers'], 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, row_data in enumerate(output_structure['data'], 2):
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Set column widths
            column_widths = [15, 15, 25, 40, 12, 15, 8, 20, 8, 12, 15]
            for col_idx, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = width
            
            # Save file
            workbook.save(output_path)
            
            self.logger.info(f'✅ Batch written to Excel: {len(records)} new records, total: {len(output_structure["data"])}')
            
        except Exception as error:
            self.logger.error(f'❌ Error writing batch to Excel: {error}')
            
            if retry_count < max_retries:
                self.logger.info(f'🔄 Retrying write in {retry_delay}ms (attempt {retry_count + 1}/{max_retries + 1})')
                await asyncio.sleep(retry_delay / 1000.0)
                return await self.write_batch_to_excel(records, retry_count + 1)
            
            raise error
        finally:
            self.is_writing = False
    
    def is_header_row(self, row: List[Any]) -> bool:
        """Check if row is a header row"""
        expected_headers = ['Số Điện Thoại', 'Số CCCD', 'Họ và Tên']
        return any(any(str(cell).find(header) != -1 for cell in row) for header in expected_headers)
    
    async def write_to_excel(self, input_record: Dict[str, Any], bhxh_result: Dict[str, Any], 
                           processing_time: int):
        """Write single record (legacy compatibility)"""
        return await self.add_to_batch(input_record, bhxh_result, processing_time)
    
    async def finalize_excel(self):
        """Finalize Excel writing (flush remaining records)"""
        try:
            await self.flush_batch(force=True)  # Force flush remaining records
            self.logger.info('📁 Excel file finalized successfully')
        except Exception as error:
            self.logger.error(f'❌ Error finalizing Excel: {error}')
            raise error
    
    async def create_backup(self) -> Optional[str]:
        """Create backup of output file"""
        try:
            output_path = self.config.files['output_excel']
            
            if not os.path.exists(output_path):
                return None
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = output_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
            
            shutil.copy2(output_path, backup_path)
            
            self.logger.info(f'💾 Backup created: {backup_path}')
            return backup_path
            
        except Exception as error:
            self.logger.error(f'❌ Error creating backup: {error}')
            return None
    
    def validate_excel_format(self, file_path: str) -> Dict[str, Any]:
        """Validate Excel file format"""
        try:
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            # Read first few rows to check headers
            headers = []
            for row in worksheet.iter_rows(max_row=1, values_only=True):
                headers = [str(cell).lower().strip() for cell in row if cell]
                break
            
            if not headers:
                return {'valid': False, 'error': 'File Excel rỗng'}
            
            # Flexible header matching - support multiple variants
            header_mappings = {
                'số cccd': ['số cccd', 'cccd', 'số cmnd', 'cmnd', 'cmt'],
                'họ và tên': ['họ và tên', 'họ tên', 'tên', 'họ và tên ']
            }
            
            missing_headers = []
            
            for required_header, variants in header_mappings.items():
                found = any(
                    any(variant in header or header in variant for header in headers)
                    for variant in variants
                )
                
                if not found:
                    missing_headers.append(required_header)
            
            if missing_headers:
                return {
                    'valid': False,
                    'error': f'Thiếu các cột bắt buộc: {", ".join(missing_headers)}'
                }
            
            return {'valid': True}
            
        except Exception as error:
            return {'valid': False, 'error': str(error)}
    
    def get_file_stats(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Get file statistics"""
        try:
            if not os.path.exists(file_path):
                return None
            
            stats = os.stat(file_path)
            workbook = load_workbook(file_path)
            
            total_records = 0
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                total_records += worksheet.max_row - 1  # Subtract header row
            
            return {
                'size': stats.st_size,
                'size_in_mb': f'{stats.st_size / (1024 * 1024):.2f}',
                'created': datetime.fromtimestamp(stats.st_ctime),
                'modified': datetime.fromtimestamp(stats.st_mtime),
                'sheets': len(workbook.sheetnames),
                'total_records': total_records
            }
            
        except Exception as error:
            self.logger.error(f'Error getting file stats: {error}')
            return None


# Singleton instance
_instance: Optional[ExcelService] = None


def get_excel_service() -> ExcelService:
    """Get Excel service instance"""
    global _instance
    if _instance is None:
        _instance = ExcelService()
    return _instance