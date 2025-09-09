#!/usr/bin/env python3
"""
Feature-6: Excel Exporter
Export dữ liệu ra file Excel
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Excel Exporter cho Feature-6"""

    @staticmethod
    def _mask_gender(gender):
        """
        Mask or anonymize gender value for privacy.
        Returns a generic label for known values, and 'Other' otherwise.
        """
        gender_map = {
            "male": "Group A",
            "female": "Group B",
            "other": "Group C",
            "unknown": "Group C"
        }
        return gender_map.get(str(gender).lower(), "Group C")
    
    def __init__(self, config: Dict[str, Any]):
        """Khởi tạo exporter"""
        self.config = config
        self.output_file = config.get('output_file', 'output.xlsx')
        self.output_sheet = config.get('output_sheet', 'Result')
        
        logger.info("✅ Excel Exporter initialized")
        logger.info(f"📊 Output file: {self.output_file}")
        logger.info(f"📋 Output sheet: {self.output_sheet}")
    
    def create_empty_data(self) -> List[Dict[str, Any]]:
        """Tạo dữ liệu trống (chỉ sử dụng khi không có dữ liệu thực)"""
        empty_data = []
        return empty_data
    
    def format_excel(self, workbook: Workbook, worksheet):
        """Format Excel file"""
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders to all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def export_to_excel(self, data: List[Dict[str, Any]] = None) -> bool:
        """Export dữ liệu ra Excel"""
        try:
            logger.info(f"📊 Exporting data to Excel: {self.output_file}")
            
            # Sử dụng dữ liệu trống nếu không có dữ liệu thực
            if data is None:
                data = self.create_empty_data()
                logger.info("📋 No data provided for export")
            
            # Tạo DataFrame
            df = pd.DataFrame(data)
            
            # Tạo workbook và worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = self.output_sheet
            
            # Thêm dữ liệu vào worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
            
            # Format Excel
            self.format_excel(workbook, worksheet)
            
            # Thêm thông tin metadata
            worksheet.cell(row=len(data) + 3, column=1, value="Thông tin xuất file:")
            worksheet.cell(row=len(data) + 4, column=1, value=f"Ngày xuất: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            worksheet.cell(row=len(data) + 5, column=1, value=f"Số bản ghi: {len(data)}")
            worksheet.cell(row=len(data) + 6, column=1, value="Hệ thống: BHXH Information System")
            
            # Lưu file
            workbook.save(self.output_file)
            
            logger.info(f"✅ Excel export completed: {self.output_file}")
            logger.info(f"📊 Records exported: {len(data)}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error exporting to Excel: {e}")
            return False
    
    def export_cccd_data(self, cccd_data: List[Dict[str, Any]]) -> bool:
        """Export dữ liệu CCCD ra Excel"""
        try:
            logger.info(f"📊 Exporting CCCD data to Excel: {self.output_file}")
            
            # Tạo DataFrame từ dữ liệu CCCD
            df = pd.DataFrame(cccd_data)
            
            # Tạo workbook và worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "CCCD Data"
            
            # Thêm dữ liệu vào worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
            
            # Format Excel
            self.format_excel(workbook, worksheet)
            
            # Lưu file
            workbook.save(self.output_file)
            
            logger.info(f"✅ CCCD Excel export completed: {self.output_file}")
            logger.info(f"📊 CCCD records exported: {len(cccd_data)}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error exporting CCCD data to Excel: {e}")
            return False
    
    def create_summary_report(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Tạo báo cáo tổng kết"""
        if not data:
            return {}
        
        df = pd.DataFrame(data)
        
        summary = {
            'total_records': len(data),
            'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'columns': list(df.columns),
            'data_types': df.dtypes.to_dict(),
            'null_counts': df.isnull().sum().to_dict(),
            'unique_counts': df.nunique().to_dict()
        }
        
        # Thống kê theo giới tính nếu có
        if 'gender' in df.columns:
            summary['gender_distribution'] = df['gender'].value_counts().to_dict()
        
        # Thống kê theo tỉnh nếu có
        if 'province' in df.columns:
            summary['province_distribution'] = df['province'].value_counts().to_dict()
        
        logger.info("📊 Summary report created")
        return summary
    
    def save_summary_report(self, summary: Dict[str, Any], filename: str = "summary_report.txt"):
        """Lưu báo cáo tổng kết"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("BÁO CÁO TỔNG KẾT XUẤT DỮ LIỆU\n")
                f.write("=" * 50 + "\n\n")
                
                f.write(f"Ngày xuất: {summary.get('export_date', 'N/A')}\n")
                f.write(f"Tổng số bản ghi: {summary.get('total_records', 0)}\n")
                f.write(f"Số cột dữ liệu: {len(summary.get('columns', []))}\n\n")
                
                f.write("Các cột dữ liệu:\n")
                for i, col in enumerate(summary.get('columns', []), 1):
                    f.write(f"  {i}. {col}\n")
                
                f.write("\nThống kê null values:\n")
                for col, count in summary.get('null_counts', {}).items():
                    f.write(f"  {col}: {count}\n")
                
                f.write("\nThống kê unique values:\n")
                for col, count in summary.get('unique_counts', {}).items():
                    f.write(f"  {col}: {count}\n")
                
                if 'gender_distribution' in summary:
                    f.write("\nPhân bố theo giới tính:\n")
                    for gender, count in summary['gender_distribution'].items():
                        masked_gender = self._mask_gender(gender)
                        f.write(f"  {masked_gender}: {count}\n")
                
                if 'province_distribution' in summary:
                    f.write("\nPhân bố theo tỉnh:\n")
                    for province, count in summary['province_distribution'].items():
                        f.write(f"  {province}: {count}\n")
            
            logger.info(f"💾 Summary report saved: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error saving summary report: {e}")
            return False

def main():
    """Test function"""
    config = {
        'output_file': 'test_output.xlsx',
        'output_sheet': 'Test Result'
    }
    
    exporter = ExcelExporter(config)
    
    # Test export với dữ liệu trống
    success = exporter.export_to_excel()
    
    if success:
        print("✅ Excel export test completed successfully")
    else:
        print("❌ Excel export test failed")

if __name__ == "__main__":
    main()